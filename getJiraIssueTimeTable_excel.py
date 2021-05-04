# -*- coding: utf-8 -*- 

from tkinter import filedialog
import openpyxl
import requests
import json
import datetime
import numpy as np

headers = {'Authorization': '[your authorization key]]'}
wantedlab_jira_domain = 'https://[your jira domain]/'
jira_api =  wantedlab_jira_domain + 'rest/internal/2/issue/'

try:
    # Open an excel file
    filename = filedialog.askopenfilename(initialdir = "C:/", title="Choose your file") 
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active

    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row

    loc_key = -1
    loc_created_time = -1
    loc_subtask = -1
    loc_fixedVersion = -1
    loc_label = -1

    for i in range(1, max_column+1):
        if sheet_obj.cell(row=1, column=i).value == 'key':
            loc_key = i
        elif sheet_obj.cell(row=1, column=i).value == 'created':
            loc_created_time = i
        elif sheet_obj.cell(row=1, column=i).value == 'fixVersion':
            loc_fixedVersion = i
        elif sheet_obj.cell(row=1, column=i).value == 'subtasks':
            loc_subtask = i
        elif sheet_obj.cell(row=1, column=i).value == 'labels':
            loc_label = i            
    
    if loc_key != -1:
        col_num_of_task = max_column + 1
        col_project_name = max_column + 2
        col_in_progress = max_column + 3
        col_in_progress_adj = max_column + 4
        col_in_review = max_column + 5
        col_resolved = max_column + 6
        col_closed = max_column + 7
        col_released = max_column + 8
        col_days_in_backlog = max_column + 9
        col_days_in_progress = max_column + 10 
        col_days_in_review = max_column + 11
        col_days_in_resolved = max_column + 12
        col_days_from_in_progress_to_resolved = max_column + 13
        col_cycletime = max_column + 14

        sheet_obj.cell(row=1, column=col_num_of_task).value = 'Num of Task'
        sheet_obj.cell(row=1, column=col_project_name).value = 'Project Name'
        sheet_obj.cell(row=1, column=col_in_progress).value = 'Date in progress'
        sheet_obj.cell(row=1, column=col_in_progress_adj).value = 'Date in progress_adjusted'
        sheet_obj.cell(row=1, column=col_in_review).value = 'Date in review'
        sheet_obj.cell(row=1, column=col_resolved).value = 'Date resolved'
        sheet_obj.cell(row=1, column=col_closed).value = 'Date closed'
        sheet_obj.cell(row=1, column=col_released).value = 'Date released'
        sheet_obj.cell(row=1, column=col_days_in_backlog).value = 'Days in Backlog'
        sheet_obj.cell(row=1, column=col_days_in_progress).value = 'Days in Progress'
        sheet_obj.cell(row=1, column=col_days_in_review).value = 'Days in Review'
        sheet_obj.cell(row=1, column=col_days_in_resolved).value = 'Days in Resolved'
        sheet_obj.cell(row=1, column=col_days_from_in_progress_to_resolved).value = 'Dev Time'
        sheet_obj.cell(row=1, column=col_cycletime).value = 'Cycle Time'
        
        for i in range(2, max_row+1):
            date_created = ''
            date_in_progress = ''
            date_in_progress_adj = ''
            date_in_review = ''
            date_resolved = ''
            date_closed = ''
            date_released = ''
            num_of_task = -1
            qa_effort = ''
            project_name = ''

            # extract project name from labes (naming convention=all capital letters)
            if loc_label != -1 and sheet_obj.cell(row=i, column=loc_label).value is not None:
                labels = sheet_obj.cell(row=i, column=loc_label).value.split(';')
                for label in labels:
                    if label.isupper() and project_name == '':
                        project_name = label                 

                sheet_obj.cell(row = i, column = col_project_name).value = project_name

            # count the number of sub-tasks
            if loc_subtask != -1 and sheet_obj.cell(row=i, column=loc_subtask).value is not None:
                subTask = sheet_obj.cell(row=i, column=loc_subtask).value.split(';')
                num_of_task = len(subTask)
                sheet_obj.cell(row = i, column = col_num_of_task).value = num_of_task

            # get issue's crated datetime
            if loc_created_time != -1:
                date_created = sheet_obj.cell(row=i, column=loc_created_time).value
            

            # get Released datetime from fixVersion
            if loc_fixedVersion != -1 and sheet_obj.cell(row=i, column=loc_fixedVersion).value is not None:
                fixedVersion = sheet_obj.cell(row=i, column=loc_fixedVersion).value.split(';')
                for ver in fixedVersion:
                    if ver[-8:].isnumeric():
                        date_released = datetime.datetime.strptime(ver[-8:], '%Y%m%d')
                        sheet_obj.cell(row=i, column=col_released).value=date_released
                        break

            # get issue's In Progress, In Review, Resolved, Closed datetime from History
            req_url = jira_api + sheet_obj.cell(row=i, column=2).value + '/activityfeed?startAt=0'
            print(req_url)
            response = requests.request('GET', req_url, headers=headers)
            
            data=json.loads(response.text.encode('utf-8'))
            history=data['items']
            isInitiated=0
            for item in history:
                if 'fieldId' in item.keys() and item['fieldId']=='status':
                    if item['to']['displayValue'] == 'In Progress' and isInitiated == 0:
                        date_in_progress = datetime.datetime.fromtimestamp(round(item['timestamp']/1000))
                        date_in_progress_adj = date_in_progress
                        sheet_obj.cell(row = i, column = col_in_progress).value = date_in_progress
                        isInitiated = 1
                    elif item['to']['displayValue'] == 'In Review':
                        date_in_review = datetime.datetime.fromtimestamp(round(item['timestamp']/1000))
                        sheet_obj.cell(row = i, column = col_in_review).value = date_in_review
                    elif item['to']['displayValue'] == 'Resolved':
                        date_resolved = datetime.datetime.fromtimestamp(round(item['timestamp']/1000))
                        sheet_obj.cell(row = i, column = col_resolved).value = date_resolved
                    elif item['to']['displayValue'] == 'Closed':
                        date_closed = datetime.datetime.fromtimestamp(round(item['timestamp']/1000))
                        sheet_obj.cell(row = i, column = col_closed).value = date_closed

            # get the status hitory of the sub-tasks of the issues 
            if num_of_task > 0:
                for st in subTask:
                    task_in_progress = list()

                    req_url_summary = wantedlab_jira_domain + 'rest/api/2/issue/' + st +'?fields=summary'
                    response_summary = requests.request('GET', req_url_summary, headers=headers)

                    summary=json.loads(response_summary.text.encode('utf-8'))['fields']['summary']

                    req_url = jira_api + st + '/activityfeed?startAt=0'
                    response_task = requests.request('GET', req_url, headers=headers)

                    data_task=json.loads(response_task.text.encode('utf-8'))
                    history_task=data_task['items']
                    for item in history_task:
                        if 'fieldId' in item.keys() and item['fieldId']=='status':
                            if item['to']['displayValue'] == 'In Progress':
                                task_in_progress.append(datetime.datetime.fromtimestamp(round(item['timestamp']/1000)))
                    
                    max_task_in_progress = max(task_in_progress) if len(task_in_progress)>0 else ''
                    
                    if max_task_in_progress != '' and (date_in_progress_adj == '' or date_in_progress_adj > max_task_in_progress):
                        date_in_progress_adj = max_task_in_progress            
            
            sheet_obj.cell(row=i, column=col_in_progress_adj).value=date_in_progress_adj

            # measure the days spending 
            #days_in_backlog    
            if date_in_progress_adj != '' and date_created != '':
                sheet_obj.cell(row = i, column = col_days_in_backlog).value = np.busday_count(date_created.date(), date_in_progress_adj.date())
            #days_in_Progress
            if (date_in_review != '' or date_resolved != '' or date_closed != '') and date_in_progress_adj != '':
                sheet_obj.cell(row = i, column = col_days_in_progress).value = np.busday_count(date_in_progress_adj.date(), date_in_review.date() if date_in_review != '' else (date_resolved.date() if date_resolved != '' else date_closed.date()))
            #days_in_review     
            if (date_resolved != '' or date_closed != '') and date_in_review != '':
                sheet_obj.cell(row = i, column = col_days_in_review).value = np.busday_count(date_in_review.date(), date_resolved.date() if date_resolved != '' else date_closed.date())
            #days_in_resolved
            if date_closed != '' and date_resolved != '':
                sheet_obj.cell(row = i, column = col_days_in_resolved).value = np.busday_count(date_resolved.date(), date_closed.date())
            #dev time : days from in progress to resolved
            if date_resolved != '' and date_in_progress_adj != '':
                if date_released != '' and date_resolved > date_released:
                    sheet_obj.cell(row = i, column = col_days_from_in_progress_to_resolved).value = np.busday_count(date_in_progress_adj.date(), date_released.date())                
                else:
                    sheet_obj.cell(row = i, column = col_days_from_in_progress_to_resolved).value = np.busday_count(date_in_progress_adj.date(), date_resolved.date())
            #cycle time
            if date_released != '' and date_in_progress_adj != '':
                sheet_obj.cell(row = i, column = col_cycletime).value = np.busday_count(date_in_progress_adj.date(), date_released.date())          
        wb_obj.save(filename)
    else:
        print("Required 'key' field is missing in the excel file.")
except Exception as e:
    print("Fail to read the file.")
    print(e)